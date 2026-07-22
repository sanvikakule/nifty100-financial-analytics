import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

INPUT_FILE = "outputs/peer_comparison.csv"
OUTPUT_FILE = "outputs/peer_comparison.xlsx"


def auto_width(ws):
    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 3


def apply_colors(ws):

    header_fill = PatternFill("solid", fgColor="1F4E78")
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red = PatternFill("solid", fgColor="F4CCCC")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    headers = [c.value for c in ws[1]]

    if "percentile" not in headers:
        return

    pcol = headers.index("percentile") + 1

    for row in range(2, ws.max_row + 1):

        val = ws.cell(row, pcol).value

        if val is None:
            continue

        if val >= 75:
            ws.cell(row, pcol).fill = green

        elif val >= 50:
            ws.cell(row, pcol).fill = yellow

        else:
            ws.cell(row, pcol).fill = red

    auto_width(ws)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def generate_report():

    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(INPUT_FILE)

    df = pd.read_csv(INPUT_FILE)

    with pd.ExcelWriter(
            OUTPUT_FILE,
            engine="openpyxl"
    ) as writer:

        for peer in sorted(df.peer_group_name.unique()):

            sheet = df[df.peer_group_name == peer]

            name = peer[:31]

            sheet.to_excel(
                writer,
                sheet_name=name,
                index=False
            )

    wb = load_workbook(OUTPUT_FILE)

    for ws in wb.worksheets:
        apply_colors(ws)

    wb.save(OUTPUT_FILE)

    print("\nExcel Report Generated!")
    print(OUTPUT_FILE)


if __name__ == "__main__":
    generate_report()