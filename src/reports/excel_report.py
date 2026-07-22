from pathlib import Path
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.screener.ranking import top_companies, bottom_companies


def format_sheet(ws):
    """Apply professional formatting to an Excel worksheet."""

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin = Side(style="thin")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = min(length + 3, 35)


def export_excel_report(filtered_df):

    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(exist_ok=True)

    output_file = output_dir / "Financial_Screener_Report.xlsx"

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        dashboard = pd.DataFrame({
            "Metric": [
                "Total Companies",
                "Highest Score",
                "Average Score",
                "Lowest Score"
            ],
            "Value": [
                len(filtered_df),
                filtered_df["composite_quality_score"].max(),
                round(filtered_df["composite_quality_score"].mean(), 2),
                filtered_df["composite_quality_score"].min()
            ]
        })

        dashboard.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

        filtered_df.to_excel(
            writer,
            sheet_name="Filter Results",
            index=False
        )

        top_companies(filtered_df).to_excel(
            writer,
            sheet_name="Top 10",
            index=False
        )

        bottom_companies(filtered_df).to_excel(
            writer,
            sheet_name="Bottom 10",
            index=False
        )

        filtered_df.sort_values(
            "quality_rank"
        ).to_excel(
            writer,
            sheet_name="Rankings",
            index=False
        )

        workbook = writer.book

        for sheet in workbook.sheetnames:
            format_sheet(workbook[sheet])

    print("\n✅ Excel Report Generated Successfully!")
    print(output_file)